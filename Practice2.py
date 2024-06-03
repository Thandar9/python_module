from openpyxl import Workbook, load_workbook

wb= load_workbook("test.xlsx")
ws = wb.active
ws = wb["Sheet1"]
#count=1
#data = iter(range(1, 17))
#for row in ws.iter_rows(min_row=2, max_row=5, min_col=2, max_col=5):
#   for cell in row:
#     cell.value =count
#     count+=1


#ws['B8'] = "平均"
#ws['B9'] = "=AVERAGE(B2:E5)"

#ws['C8'] = "合計"
#ws['C9'] = "=SUM(B2:E5)"

input_ws=wb.create_sheet("DataInput")

for i in range(1,11):
    input_ws.cell(i,1,i)

for row in range(1,11):
    cell_data=input_ws.cell(i,1).value
    ws.cell(i+1,6,cell_data+10)

wb.save("test.xlsx")