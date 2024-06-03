from openpyxl import Workbook, load_workbook
#wb=Workbook()
#wb=load_workbook()

wb= load_workbook("test.xlsx")
ws = wb.active
ws = wb["Sheet1"]

#print(ws['A1'].value)
#print(ws['B1'].value)
#for row in ws.rows:
 #   for cell in row:
  #    print(cell.value)

for row in ws.iter_rows(min_row=1, min_col=1, max_row=8, max_col=10):
    for cell in row:
        print(cell.value)
    
#ws['A1'] = 1
#ws.cell(1, 1, 1)
#ws['B1'] = 2
#ws.cell(1, 2, 2)
#ws['A2'] = 3
#ws.cell(2, 1, 3)
#ws['B2'] = 4
#ws.cell(2, 2, 4)
#ws.append([1,2,3,4])

wb.save("test.xlsx")